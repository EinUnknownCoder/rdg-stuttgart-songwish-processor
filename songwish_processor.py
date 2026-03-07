#!/usr/bin/env python3
"""
RDG Stuttgart Song Wish Processor

Validates song wish requests and generates output Excel with messages and songlist.
"""

import re
import unicodedata
from urllib.parse import urlparse, parse_qs, urlencode, quote
import pandas as pd
import yt_dlp
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Configuration
FORM_URL = "https://forms.gle/KTg2MvaRo8TFVK7q7"  # Replace with actual form URL
MAX_SONG_DURATION_SECONDS = 90
FIRST_GUARANTEED_COUNT = 50

# Blocked songs list file
BLOCKED_SONGS_FILE = "blocked_songs.xlsx"


def normalize_text(text):
    """Normalize text by removing special chars, spaces, and converting to lowercase."""
    if pd.isna(text) or text is None:
        return ""
    text = str(text).lower()
    # Remove accents
    text = unicodedata.normalize('NFKD', text).encode('ASCII', 'ignore').decode('ASCII')
    # Remove all non-alphanumeric characters
    text = re.sub(r'[^a-z0-9]', '', text)
    return text


def clean_youtube_url(url):
    """Remove playlist parameters from YouTube URL."""
    if pd.isna(url) or not url:
        return None
    url = str(url).strip()

    # Parse the URL
    parsed = urlparse(url)
    query_params = parse_qs(parsed.query)

    # Remove list and index parameters
    params_to_remove = ['list', 'index', 'start_radio', 'pp']
    for param in params_to_remove:
        query_params.pop(param, None)

    # Rebuild URL
    if query_params:
        new_query = urlencode(query_params, doseq=True)
        clean_url = f"{parsed.scheme}://{parsed.netloc}{parsed.path}?{new_query}"
    else:
        clean_url = f"{parsed.scheme}://{parsed.netloc}{parsed.path}"

    return clean_url


def parse_timestamp(ts):
    """Parse timestamp string to seconds. Handles MM:SS:00 and MM:SS formats."""
    if pd.isna(ts) or not ts:
        return 0

    ts_str = str(ts).strip()

    # Handle time format from Excel (HH:MM:SS or MM:SS:00)
    parts = ts_str.split(':')

    try:
        if len(parts) == 3:
            # Could be HH:MM:SS or MM:SS:00
            # If first part is 00 or small number and last is 00, it's likely MM:SS:00
            h, m, s = int(parts[0]), int(parts[1]), int(parts[2])
            if h < 10 and s == 0:
                # Likely MM:SS:00 format (minutes:seconds:00)
                return h * 60 + m
            else:
                # Standard HH:MM:SS
                return h * 3600 + m * 60 + s
        elif len(parts) == 2:
            # MM:SS format
            return int(parts[0]) * 60 + int(parts[1])
        else:
            return int(float(ts_str))
    except (ValueError, TypeError):
        return 0


def get_youtube_info(url):
    """Fetch YouTube video information using yt-dlp."""
    if not url:
        return None

    ydl_opts = {
        'quiet': True,
        'no_warnings': True,
        'extract_flat': False,
        'skip_download': True,
        'socket_timeout': 30,  # 30 second timeout for network operations
        'source_address': '0.0.0.0',  # Force IPv4 to avoid YouTube connection issues
    }

    try:
        with yt_dlp.YoutubeDL(ydl_opts) as ydl:
            info = ydl.extract_info(url, download=False)
            return {
                'title': info.get('title', ''),
                'description': info.get('description', ''),
                'duration': info.get('duration', 0),
                'age_limit': info.get('age_limit', 0),
                'categories': info.get('categories', []),
                'tags': info.get('tags', []),
                'channel': info.get('channel', ''),
                'uploader': info.get('uploader', ''),
            }
    except Exception as e:
        return {'error': str(e)}


def check_is_lyric_video(video_info):
    """Check if the video appears to be a lyric video."""
    if not video_info or 'error' in video_info:
        return False, "Video konnte nicht abgerufen werden / Could not fetch video"

    title = (video_info.get('title') or '').lower()
    description = (video_info.get('description') or '').lower()
    tags = [t.lower() for t in video_info.get('tags', [])] if video_info.get('tags') else []

    # Positive indicators for lyric video
    lyric_indicators = ['lyric', 'lyrics', 'lyric video', 'lyrics video', 'letra', 'text', 'sing-along', 'singalong']

    # Negative indicators (official MV, dance practice, etc.)
    negative_indicators = ['official mv', 'official music video', 'dance practice',
                          'dance practice video', 'choreography video', 'performance video',
                          'm/v', '(mv)', '[mv]']

    title_lower = title.lower()

    # Check for negative indicators first
    for neg in negative_indicators:
        if neg in title_lower:
            return False, f"Kein Lyric Video ('{neg}' im Titel gefunden) / Not a lyric video ('{neg}' found in title)"

    # Check for positive indicators
    for pos in lyric_indicators:
        if pos in title_lower or pos in description or pos in tags:
            return True, None

    # If no clear indicator, we'll accept it but note it's uncertain
    return True, None  # Accept by default, but could be flagged for manual review


def check_artist_title_match(video_info, artist, song_title):
    """Check if artist and title are present in YouTube video title."""
    if not video_info or 'error' in video_info:
        return False, "Video konnte nicht abgerufen werden / Could not fetch video"

    yt_title = video_info.get('title', '')
    yt_title_normalized = normalize_text(yt_title)

    errors = []

    # Check artist
    artist_normalized = normalize_text(artist)
    if artist_normalized and artist_normalized not in yt_title_normalized:
        errors.append(f"Künstler '{artist}' nicht im YouTube-Titel gefunden / Artist '{artist}' not found in YouTube title")

    # Check song title
    title_normalized = normalize_text(song_title)
    if title_normalized and title_normalized not in yt_title_normalized:
        errors.append(f"Songtitel '{song_title}' nicht im YouTube-Titel gefunden / Song title '{song_title}' not found in YouTube title")

    if errors:
        return False, "; ".join(errors)
    return True, None


def check_duration(start_ts, end_ts):
    """Check if the song section is within the allowed duration."""
    start_seconds = parse_timestamp(start_ts)
    end_seconds = parse_timestamp(end_ts)

    duration = end_seconds - start_seconds

    if duration <= 0:
        return False, f"Ungültige Timestamps (Start: {start_ts}, Ende: {end_ts}) / Invalid timestamps (Start: {start_ts}, End: {end_ts})"

    if duration > MAX_SONG_DURATION_SECONDS:
        return False, f"Songabschnitt zu lang ({duration}s > {MAX_SONG_DURATION_SECONDS}s) / Song section too long ({duration}s > {MAX_SONG_DURATION_SECONDS}s)"

    return True, None


def check_age_restriction(video_info):
    """Check if the video is age-restricted (18+)."""
    if not video_info or 'error' in video_info:
        return True, None  # Can't check, assume OK

    age_limit = video_info.get('age_limit', 0)
    if age_limit and age_limit >= 18:
        return False, "18+ Video nicht erlaubt / 18+ video not allowed"

    return True, None


def load_blocked_songs():
    """Load blocked songs list from Excel file."""
    try:
        df = pd.read_excel(BLOCKED_SONGS_FILE)
        blocked = {}
        for _, row in df.iterrows():
            artist = normalize_text(row.get('Artist', ''))
            title = normalize_text(row.get('Title', ''))
            grund = row.get('Grund', '')
            if artist and title:
                blocked[(artist, title)] = grund if pd.notna(grund) else ''
        return blocked
    except FileNotFoundError:
        return {}


def check_blocked_song(artist, title, blocked_songs):
    """Check if the song is in the blocked list."""
    artist_norm = normalize_text(artist)
    title_norm = normalize_text(title)

    if (artist_norm, title_norm) in blocked_songs:
        return False, f"Das Lied befindet sich auf der Liste der gesperrten Songs (z.B. weil es 18+ ist) / The song is on the list of blocked songs (e.g. because it is 18+)"

    return True, None


def validate_song(url, artist, title, start_ts, end_ts, blocked_songs):
    """Validate a single song and return errors if any."""
    errors = []

    # Clean URL
    clean_url = clean_youtube_url(url)
    if not clean_url:
        return ["Keine URL angegeben / No URL provided"], clean_url

    # Fetch video info
    video_info = get_youtube_info(clean_url)

    if video_info and 'error' in video_info:
        errors.append(f"YouTube-Fehler: {video_info['error']} / YouTube error: {video_info['error']}")
        return errors, clean_url

    # Check artist and title match
    match_ok, match_error = check_artist_title_match(video_info, artist, title)
    if not match_ok:
        errors.append(match_error)

    # Check if lyric video
    lyric_ok, lyric_error = check_is_lyric_video(video_info)
    if not lyric_ok:
        errors.append(lyric_error)

    # Check duration
    duration_ok, duration_error = check_duration(start_ts, end_ts)
    if not duration_ok:
        errors.append(duration_error)

    # Check age restriction
    age_ok, age_error = check_age_restriction(video_info)
    if not age_ok:
        errors.append(age_error)

    # Check blocked songs
    blocked_ok, blocked_error = check_blocked_song(artist, title, blocked_songs)
    if not blocked_ok:
        errors.append(blocked_error)

    return errors, clean_url


def create_contact_url(row):
    """Create contact URL based on preferred communication method.

    Returns:
        tuple: (url, contact_type) where contact_type is 'instagram', 'whatsapp', or 'other'
    """
    preference = row.get('Bevorzugte Kommunikation\nPreferred communication', '')

    if 'Instagram' in str(preference):
        instagram = row.get('Instagram @Name', '')
        if pd.notna(instagram) and instagram:
            # Remove @ if present
            instagram = str(instagram).strip()
            if instagram.startswith('@'):
                instagram = instagram[1:]
            return f"https://www.instagram.com/{instagram}", 'instagram'

    elif 'WhatsApp' in str(preference):
        phone = row.get('WhatsApp Number', '')
        if pd.notna(phone) and phone:
            # Clean phone number (remove spaces, dashes, etc.)
            phone = re.sub(r'[^\d+]', '', str(phone))
            # Ensure it starts with country code
            if not phone.startswith('+'):
                phone = '+' + phone
            return f"https://wa.me/{phone.replace('+', '')}", 'whatsapp'

    # Fallback to other contact method
    other = row.get('Weitere Kontaktmöglichkeit\nFurther contact information', '')
    if pd.notna(other) and other:
        return str(other), 'other'

    return "", 'other'


def get_greeting_name(row):
    """Get name for greeting based on contact method."""
    preference = row.get('Bevorzugte Kommunikation\nPreferred communication', '')

    if 'Instagram' in str(preference):
        instagram = row.get('Instagram @Name', '')
        if pd.notna(instagram) and instagram:
            name = str(instagram).strip()
            if name.startswith('@'):
                name = name[1:]
            return name

    return None


def create_message(row, errors, language, form_url, artist, title):
    """Create message based on validation result and language."""
    name = get_greeting_name(row)
    is_german = '🇩🇪' in str(language) or 'Deutsch' in str(language)
    song_info = f"{artist} - {title}"

    if is_german:
        if name:
            greeting = f"Hallo {name}! 👋"
        else:
            greeting = "Hallo! 👋"

        if not errors:
            message = f"""{greeting}

Vielen Dank für deinen Songwunsch beim RDG Stuttgart! 🎵

Dein Songwunsch: {song_info}

Dein Songwunsch wurde erfolgreich geprüft und wird auf die Playlist aufgenommen, sobald du auf diese Nachricht antwortest/reagierst.

Wir freuen uns auf dich! 🎉"""
        else:
            error_list = "\n".join([f"• {e.split(' / ')[0]}" for e in errors])
            message = f"""{greeting}

Vielen Dank für deinen Songwunsch beim RDG Stuttgart! 🎵

Dein Songwunsch: {song_info}

Leider gibt es ein Problem mit deinem Songwunsch:

{error_list}

Bitte korrigiere deinen Songwunsch über dieses Formular: {form_url}

Antworte auf diese Nachricht, sobald du die Korrektur vorgenommen hast. Ansonsten ist der Songwunsch leider ungültig.

Bei Fragen kannst du dich gerne melden! 💬"""
    else:
        if name:
            greeting = f"Hello {name}! 👋"
        else:
            greeting = "Hello! 👋"

        if not errors:
            message = f"""{greeting}

Thank you for your song wish at RDG Stuttgart! 🎵

Your song wish: {song_info}

Your song wish has been successfully verified and will be added to the playlist once you reply/react to this message.

We look forward to seeing you! 🎉"""
        else:
            error_list = "\n".join([f"• {e.split(' / ')[-1]}" for e in errors])
            message = f"""{greeting}

Thank you for your song wish at RDG Stuttgart! 🎵

Your song wish: {song_info}

Unfortunately, there is a problem with your song wish:

{error_list}

Please correct your song wish using this form: {form_url}

Reply to this message once you have made the correction. Otherwise, your song wish will be invalid.

Feel free to reach out if you have any questions! 💬"""

    return message


def create_blocked_songs_template():
    """Create empty blocked songs Excel template if it doesn't exist."""
    try:
        pd.read_excel(BLOCKED_SONGS_FILE)
    except FileNotFoundError:
        df = pd.DataFrame(columns=['Artist', 'Title'])
        df.to_excel(BLOCKED_SONGS_FILE, index=False)
        print(f"Created blocked songs template: {BLOCKED_SONGS_FILE}")


def generate_messages_html(results, output_html_file):
    """Generate an HTML file with clickable send buttons for WhatsApp and Instagram."""
    import html as html_module

    rows_html = []
    for i, result in enumerate(results[:FIRST_GUARANTEED_COUNT]):
        has_errors = bool(result['errors1'])
        status_class = 'error' if has_errors else 'ok'
        status_text = 'Fehler' if has_errors else 'OK'
        artist = html_module.escape(str(result['artist1']) if pd.notna(result['artist1']) else '')
        title = html_module.escape(str(result['title1']) if pd.notna(result['title1']) else '')
        song = f"{artist} - {title}"

        # Requester display
        instagram = result.get('instagram', '')
        requester = ''
        if pd.notna(instagram) and instagram:
            requester = str(instagram).strip()
            if requester.startswith('@'):
                requester = requester[1:]
        if not requester:
            email = result.get('email', '')
            requester = str(email) if pd.notna(email) and email else ''
        requester = html_module.escape(requester)

        contact_type = result.get('contact_type', 'other')
        contact_url = result.get('contact_url', '')
        message = result['message']
        # Escape message for embedding in JS data attribute
        message_escaped = html_module.escape(message)

        if contact_type == 'whatsapp':
            wa_url = contact_url + '?text=' + quote(message, safe='')
            button_html = (
                f'<a href="{html_module.escape(wa_url)}" target="_blank" class="btn btn-whatsapp">'
                f'\U0001f4e9 Senden</a>'
            )
        elif contact_type == 'instagram':
            button_html = (
                f'<button class="btn btn-instagram" '
                f'data-message="{message_escaped}" '
                f'data-url="{html_module.escape(contact_url)}" '
                f'onclick="copyAndOpen(this)">'
                f'\U0001f4cb Kopieren &amp; \u00d6ffnen</button>'
            )
        else:
            contact_display = html_module.escape(contact_url) if contact_url else '-'
            button_html = f'<span class="contact-info">{contact_display}</span>'

        rows_html.append(
            f'<tr class="{status_class}">'
            f'<td>{i + 1}</td>'
            f'<td><span class="badge {status_class}">{status_text}</span></td>'
            f'<td>{requester}</td>'
            f'<td>{song}</td>'
            f'<td>{button_html}</td>'
            f'</tr>'
        )

    html_content = f'''<!DOCTYPE html>
<html lang="de">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>RDG Stuttgart - Songwish Messages</title>
<style>
  body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif; margin: 20px; background: #f5f5f5; }}
  h1 {{ color: #333; }}
  table {{ border-collapse: collapse; width: 100%; background: white; box-shadow: 0 1px 3px rgba(0,0,0,0.1); }}
  th {{ background: #444; color: white; padding: 10px 12px; text-align: left; }}
  td {{ padding: 8px 12px; border-bottom: 1px solid #eee; vertical-align: middle; }}
  tr.ok {{ background: #f0fff0; }}
  tr.error {{ background: #fff0f0; }}
  tr:hover {{ filter: brightness(0.97); }}
  .badge {{ padding: 3px 8px; border-radius: 4px; font-size: 0.85em; font-weight: bold; }}
  .badge.ok {{ background: #4caf50; color: white; }}
  .badge.error {{ background: #f44336; color: white; }}
  .btn {{ display: inline-block; padding: 6px 14px; border-radius: 6px; text-decoration: none;
          font-size: 0.9em; font-weight: 500; cursor: pointer; border: none; color: white; }}
  .btn-whatsapp {{ background: #25d366; }}
  .btn-whatsapp:hover {{ background: #1da851; }}
  .btn-instagram {{ background: #e1306c; }}
  .btn-instagram:hover {{ background: #c13584; }}
  .btn.copied {{ background: #888 !important; }}
  .contact-info {{ color: #666; font-size: 0.9em; }}
</style>
</head>
<body>
<h1>RDG Stuttgart - Songwish Messages</h1>
<p>{min(len(results), FIRST_GUARANTEED_COUNT)} Nachrichten (erste {FIRST_GUARANTEED_COUNT} garantierte Requests)</p>
<table>
<thead>
  <tr><th>#</th><th>Status</th><th>Requester</th><th>Song</th><th>Aktion</th></tr>
</thead>
<tbody>
{"".join(rows_html)}
</tbody>
</table>
<script>
function copyAndOpen(btn) {{
  var message = btn.getAttribute("data-message");
  var url = btn.getAttribute("data-url");
  navigator.clipboard.writeText(message).then(function() {{
    var original = btn.innerHTML;
    btn.innerHTML = "\\u2705 Kopiert!";
    btn.classList.add("copied");
    setTimeout(function() {{
      btn.innerHTML = original;
      btn.classList.remove("copied");
    }}, 2000);
    window.open(url, "_blank");
  }}).catch(function() {{
    // Fallback for older browsers
    var ta = document.createElement("textarea");
    ta.value = message;
    document.body.appendChild(ta);
    ta.select();
    document.execCommand("copy");
    document.body.removeChild(ta);
    btn.innerHTML = "\\u2705 Kopiert!";
    btn.classList.add("copied");
    setTimeout(function() {{
      btn.innerHTML = "\\U0001f4cb Kopieren & \\u00d6ffnen";
      btn.classList.remove("copied");
    }}, 2000);
    window.open(url, "_blank");
  }});
}}
</script>
</body>
</html>'''

    with open(output_html_file, 'w', encoding='utf-8') as f:
        f.write(html_content)
    print(f"HTML messages saved to: {output_html_file}")


def process_songwishes(input_file, output_file, form_url=FORM_URL):
    """Main processing function."""
    print(f"Reading {input_file}...")
    df = pd.read_excel(input_file)

    # Skip first row if it contains translations (no longer needed with new format)
    # if pd.notna(df.iloc[0].get('Sprache der Regeln')) and 'Language' in str(df.iloc[0].get('Sprache der Regeln')):
    #     df = df.iloc[1:].reset_index(drop=True)

    print(f"Processing {len(df)} song wishes...")

    # Load blocked songs
    blocked_songs = load_blocked_songs()
    print(f"Loaded {len(blocked_songs)} blocked songs")

    # Process each song wish
    results = []

    for idx, row in df.iterrows():
        print(f"Processing request {idx + 1}/{len(df)}...")

        # First song
        url1 = row.get('YT URL', '')
        artist1 = row.get('Künstler\nArtist', '')
        title1 = row.get('Songname\nSong Title', '')
        part1 = row.get('Teil des Liedes\nPart of the Song', '')
        start1 = row.get('Start Timestamp', '')
        end1 = row.get('End Timestamp', '')
        note1 = row.get('Anmerkung\nAdditional Information', '')

        errors1, clean_url1 = validate_song(url1, artist1, title1, start1, end1, blocked_songs)

        # Second song
        url2 = row.get('YT URL.1', '')
        artist2 = row.get('Künstler\nArtist.1', '')
        title2 = row.get('Songname\nSong Title.1', '')
        part2 = row.get('Teil des Liedes\nPart of the Song.1', '')
        start2 = row.get('Start Timestamp.1', '')
        end2 = row.get('End Timestamp.1', '')
        note2 = row.get('Anmerkung\nAdditional Information.1', '')

        errors2, clean_url2 = validate_song(url2, artist2, title2, start2, end2, blocked_songs) if pd.notna(url2) and url2 else ([], None)

        contact_url, contact_type = create_contact_url(row)

        results.append({
            'row_index': idx,
            'email': row.get('Email Address', ''),
            'language': row.get('Sprache der Regeln\nLanguage of the Rules', ''),
            'contact_pref': row.get('Bevorzugte Kommunikation\nPreferred communication', ''),
            'instagram': row.get('Instagram @Name', ''),
            'whatsapp': row.get('WhatsApp Number', ''),
            'other_contact': row.get('Weitere Kontaktmöglichkeit\nFurther contact information', ''),
            # Song 1
            'url1': clean_url1,
            'artist1': artist1,
            'title1': title1,
            'part1': part1,
            'start1': start1,
            'end1': end1,
            'note1': note1,
            'errors1': errors1,
            # Song 2
            'url2': clean_url2,
            'artist2': artist2,
            'title2': title2,
            'part2': part2,
            'start2': start2,
            'end2': end2,
            'note2': note2,
            'errors2': errors2,
            # Contact
            'contact_url': contact_url,
            'contact_type': contact_type,
            'message': create_message(row, errors1, row.get('Sprache der Regeln\nLanguage of the Rules', ''), form_url, artist1, title1),
            'ok_message': create_message(row, [], row.get('Sprache der Regeln\nLanguage of the Rules', ''), form_url, artist1, title1),
        })

    # Create output Excel
    print(f"Creating output file: {output_file}...")

    wb = Workbook()

    # Sheet 1: Messages (first 50 requests, first song only)
    ws_messages = wb.active
    ws_messages.title = "Messages"

    # Headers
    headers = ['#', 'Contact URL', 'Message', 'Status', 'Artist', 'Title', 'Errors', 'OK Message']
    for col, header in enumerate(headers, 1):
        cell = ws_messages.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Data (first 50)
    error_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    success_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

    for i, result in enumerate(results[:FIRST_GUARANTEED_COUNT]):
        row_num = i + 2
        ws_messages.cell(row=row_num, column=1, value=i + 1)
        ws_messages.cell(row=row_num, column=2, value=result['contact_url'])
        ws_messages.cell(row=row_num, column=3, value=result['message'])

        has_errors = bool(result['errors1'])
        status = "Fehler / Error" if has_errors else "OK"
        ws_messages.cell(row=row_num, column=4, value=status)
        ws_messages.cell(row=row_num, column=5, value=result['artist1'])
        ws_messages.cell(row=row_num, column=6, value=result['title1'])
        ws_messages.cell(row=row_num, column=7, value="; ".join(result['errors1']) if result['errors1'] else "")
        ws_messages.cell(row=row_num, column=8, value=result['ok_message'])

        # Color coding
        fill = error_fill if has_errors else success_fill
        for col in range(1, 9):
            ws_messages.cell(row=row_num, column=col).fill = fill

    # Adjust column widths
    ws_messages.column_dimensions['A'].width = 5
    ws_messages.column_dimensions['B'].width = 40
    ws_messages.column_dimensions['C'].width = 80
    ws_messages.column_dimensions['D'].width = 15
    ws_messages.column_dimensions['E'].width = 20
    ws_messages.column_dimensions['F'].width = 30
    ws_messages.column_dimensions['G'].width = 50
    ws_messages.column_dimensions['H'].width = 80

    # Sheet 2: Songlist
    ws_songlist = wb.create_sheet("Songlist")

    # Headers matching request.xlsx format + additional columns
    songlist_headers = [
        'YouTube-URL', 'Artist', 'Title', 'Description', 'Requester/Dancer',
        'Start: Minute', 'Start: Second', 'End: Minute', 'End: Seconds',
        'Start in Seconds', 'End in Seconds',
        '#', 'Anmerkung', 'Errors',
        'Category', 'Duration', 'Artist CAPS', 'Title CAPS', 'Timestamp'
    ]

    for col, header in enumerate(songlist_headers, 1):
        cell = ws_songlist.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Add first songs first, then second songs
    row_num = 2
    song_counter = 1

    # First songs (all)
    for result in results:
        if not result['url1']:
            continue

        start_seconds = parse_timestamp(result['start1'])
        end_seconds = parse_timestamp(result['end1'])

        ws_songlist.cell(row=row_num, column=1, value=result['url1'])
        ws_songlist.cell(row=row_num, column=2, value=result['artist1'])
        ws_songlist.cell(row=row_num, column=3, value=result['title1'])
        # Description (Teil des Liedes) - clean up "Other" placeholder
        part1_value = result['part1'] if pd.notna(result['part1']) else ""
        if 'Other (Please use "Additional Information" Text Field)' in str(part1_value):
            part1_value = str(part1_value).replace('Other (Please use "Additional Information" Text Field)', '').strip()
        if not part1_value:
            part1_value = "Chorus"
        ws_songlist.cell(row=row_num, column=4, value=part1_value)
        # Use Instagram name if available, otherwise email
        requester = result['instagram'] if pd.notna(result['instagram']) and result['instagram'] else result['email']
        if requester and str(requester).startswith('@'):
            requester = str(requester)[1:]
        ws_songlist.cell(row=row_num, column=5, value=requester)  # Requester
        ws_songlist.cell(row=row_num, column=6, value=start_seconds // 60)  # Start minute
        ws_songlist.cell(row=row_num, column=7, value=start_seconds % 60)  # Start second
        ws_songlist.cell(row=row_num, column=8, value=end_seconds // 60)  # End minute
        ws_songlist.cell(row=row_num, column=9, value=end_seconds % 60)  # End second
        ws_songlist.cell(row=row_num, column=10).value = f'=F{row_num}*60+G{row_num}'
        ws_songlist.cell(row=row_num, column=11).value = f'=H{row_num}*60+I{row_num}'
        ws_songlist.cell(row=row_num, column=12, value=song_counter)
        ws_songlist.cell(row=row_num, column=13, value=result['note1'] if pd.notna(result['note1']) else "")
        ws_songlist.cell(row=row_num, column=14, value="; ".join(result['errors1']) if result['errors1'] else "")
        # Category: Top 50 for first guaranteed songs, Pool for rest
        ws_songlist.cell(row=row_num, column=15, value="Top 50" if song_counter <= FIRST_GUARANTEED_COUNT else "Pool")
        # Duration formula: End - Start + 10
        ws_songlist.cell(row=row_num, column=16).value = f'=K{row_num}-J{row_num}+10'
        # Artist CAPS formula
        ws_songlist.cell(row=row_num, column=17).value = f'=UPPER(B{row_num})'
        # Title CAPS formula
        ws_songlist.cell(row=row_num, column=18).value = f'=UPPER(C{row_num})'
        # Timestamp formula: m:ss - m:ss
        ws_songlist.cell(row=row_num, column=19).value = f'=INT(J{row_num}/60)&":"&TEXT(MOD(J{row_num},60),"00")&" - "&INT(K{row_num}/60)&":"&TEXT(MOD(K{row_num},60),"00")'

        # Mark errors with red background
        if result['errors1']:
            for col in range(1, 20):
                ws_songlist.cell(row=row_num, column=col).fill = error_fill

        row_num += 1
        song_counter += 1

    # Second songs
    for result in results:
        if not result['url2'] or pd.isna(result['url2']):
            continue

        start_seconds = parse_timestamp(result['start2'])
        end_seconds = parse_timestamp(result['end2'])

        ws_songlist.cell(row=row_num, column=1, value=result['url2'])
        ws_songlist.cell(row=row_num, column=2, value=result['artist2'])
        ws_songlist.cell(row=row_num, column=3, value=result['title2'])
        # Description (Teil des Liedes) - clean up "Other" placeholder
        part2_value = result['part2'] if pd.notna(result['part2']) else ""
        if 'Other (Please use "Additional Information" Text Field)' in str(part2_value):
            part2_value = str(part2_value).replace('Other (Please use "Additional Information" Text Field)', '').strip()
        if not part2_value:
            part2_value = "Chorus"
        ws_songlist.cell(row=row_num, column=4, value=part2_value)
        # Use Instagram name if available, otherwise email
        requester = result['instagram'] if pd.notna(result['instagram']) and result['instagram'] else result['email']
        if requester and str(requester).startswith('@'):
            requester = str(requester)[1:]
        ws_songlist.cell(row=row_num, column=5, value=f"Pool: {requester}")  # Requester (second wish)
        ws_songlist.cell(row=row_num, column=6, value=start_seconds // 60)  # Start minute
        ws_songlist.cell(row=row_num, column=7, value=start_seconds % 60)  # Start second
        ws_songlist.cell(row=row_num, column=8, value=end_seconds // 60)  # End minute
        ws_songlist.cell(row=row_num, column=9, value=end_seconds % 60)  # End second
        ws_songlist.cell(row=row_num, column=10).value = f'=F{row_num}*60+G{row_num}'
        ws_songlist.cell(row=row_num, column=11).value = f'=H{row_num}*60+I{row_num}'
        ws_songlist.cell(row=row_num, column=12, value=song_counter)
        ws_songlist.cell(row=row_num, column=13, value=result['note2'] if pd.notna(result['note2']) else "")
        ws_songlist.cell(row=row_num, column=14, value="; ".join(result['errors2']) if result['errors2'] else "")
        # Category: Second wishes are always Pool
        ws_songlist.cell(row=row_num, column=15, value="Pool")
        # Duration formula: End - Start + 10
        ws_songlist.cell(row=row_num, column=16).value = f'=K{row_num}-J{row_num}+10'
        # Artist CAPS formula
        ws_songlist.cell(row=row_num, column=17).value = f'=UPPER(B{row_num})'
        # Title CAPS formula
        ws_songlist.cell(row=row_num, column=18).value = f'=UPPER(C{row_num})'
        # Timestamp formula: m:ss - m:ss
        ws_songlist.cell(row=row_num, column=19).value = f'=INT(J{row_num}/60)&":"&TEXT(MOD(J{row_num},60),"00")&" - "&INT(K{row_num}/60)&":"&TEXT(MOD(K{row_num},60),"00")'

        # Mark errors with red background
        if result['errors2']:
            for col in range(1, 20):
                ws_songlist.cell(row=row_num, column=col).fill = error_fill

        row_num += 1
        song_counter += 1

    # Adjust column widths for Songlist
    ws_songlist.column_dimensions['A'].width = 50
    ws_songlist.column_dimensions['B'].width = 20
    ws_songlist.column_dimensions['C'].width = 30
    ws_songlist.column_dimensions['D'].width = 15
    ws_songlist.column_dimensions['E'].width = 30
    ws_songlist.column_dimensions['L'].width = 5
    ws_songlist.column_dimensions['M'].width = 40
    ws_songlist.column_dimensions['N'].width = 50
    ws_songlist.column_dimensions['O'].width = 10
    ws_songlist.column_dimensions['P'].width = 10
    ws_songlist.column_dimensions['Q'].width = 20
    ws_songlist.column_dimensions['R'].width = 30
    ws_songlist.column_dimensions['S'].width = 20

    # Save
    wb.save(output_file)
    print(f"Output saved to: {output_file}")

    # Generate HTML messages file
    html_file = output_file.rsplit('.', 1)[0] + '_messages.html'
    generate_messages_html(results, html_file)

    # Print summary
    total_errors_song1 = sum(1 for r in results if r['errors1'])
    total_errors_song2 = sum(1 for r in results if r['errors2'] and r['url2'])
    print(f"\n=== Summary ===")
    print(f"Total requests: {len(results)}")
    print(f"First songs with errors: {total_errors_song1}")
    print(f"Second songs with errors: {total_errors_song2}")
    print(f"First 50 guaranteed: {min(len(results), FIRST_GUARANTEED_COUNT)}")


if __name__ == "__main__":
    # Create blocked songs template if needed
    create_blocked_songs_template()

    # Process song wishes
    process_songwishes(
        input_file="songwish.xlsx",
        output_file="output.xlsx",
        form_url=FORM_URL
    )
